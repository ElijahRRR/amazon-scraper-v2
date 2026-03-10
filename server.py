"""
Amazon 产品采集系统 v2 - 中央服务器（FastAPI）
端口 8899
提供 API 端点和 Web UI
"""
import os
import io
import csv
import json
import random
import zipfile
import re
import asyncio
import logging
import time
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Optional, List, Dict, Any

from fastapi import FastAPI, Request, UploadFile, File, Form, Query, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import uvicorn
import openpyxl

import config
from database import Database, get_db, close_db
from models import RESULT_FIELDS

# 日志
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ==================== 生命周期 ====================

@asynccontextmanager
async def lifespan(app):
    """应用生命周期管理"""
    # startup
    db = await get_db()
    logger.info("✅ 数据库初始化完成")
    asyncio.create_task(_timeout_task_loop())
    yield
    # shutdown
    await close_db()
    logger.info("🛑 服务器关闭")


# FastAPI 应用
app = FastAPI(title="Amazon Scraper v2", version="2.0.0", lifespan=lifespan)

# 静态文件和模板
app.mount("/static", StaticFiles(directory=config.STATIC_DIR), name="static")
templates = Jinja2Templates(directory=config.TEMPLATE_DIR)

# ==================== Worker 在线状态追踪 ====================
# 存储 worker 的最后心跳时间和统计
_worker_registry: Dict[str, Dict] = {}


_WORKER_ID_RE = re.compile(r'^[\w\-]{1,64}$')

def _register_worker(worker_id: str, enable_screenshot: bool = None, ip: str = None):
    """注册/更新 worker 心跳"""
    if not _WORKER_ID_RE.match(worker_id):
        return  # 拒绝非法 worker_id
    now = time.time()
    if worker_id not in _worker_registry:
        _worker_registry[worker_id] = {
            "worker_id": worker_id,
            "first_seen": now,
            "last_seen": now,
            "tasks_pulled": 0,
            "results_submitted": 0,
            "enable_screenshot": True,
            "ip": None,
        }
    _worker_registry[worker_id]["last_seen"] = now
    if enable_screenshot is not None:
        _worker_registry[worker_id]["enable_screenshot"] = enable_screenshot
    if ip is not None:
        _worker_registry[worker_id]["ip"] = ip


# ==================== 运行时设置（可通过 API 修改，Worker 定期同步）====================

def _default_settings() -> dict:
    """从 config.py 生成默认设置（用于初始化和恢复默认）"""
    return {
        "zip_code": config.DEFAULT_ZIP_CODE,
        "max_retries": config.MAX_RETRIES,
        "request_timeout": config.REQUEST_TIMEOUT,
        "proxy_api_url": config.PROXY_API_URL_AUTH,
        "token_bucket_rate": config.TOKEN_BUCKET_RATE,
        "initial_concurrency": config.INITIAL_CONCURRENCY,
        "min_concurrency": config.MIN_CONCURRENCY,
        "max_concurrency": config.MAX_CONCURRENCY,
        "session_rotate_every": config.SESSION_ROTATE_EVERY,
        "screenshot_browsers": 1,
        "screenshot_pages_per_browser": 3,
        "adjust_interval": config.ADJUST_INTERVAL_S,
        "target_latency": config.TARGET_LATENCY_S,
        "max_latency": config.MAX_LATENCY_S,
        "target_success_rate": config.TARGET_SUCCESS_RATE,
        "min_success_rate": config.MIN_SUCCESS_RATE,
        "block_rate_threshold": config.BLOCK_RATE_THRESHOLD,
        "cooldown_after_block": 15,
        "proxy_bandwidth_mbps": config.PROXY_BANDWIDTH_MBPS,
        "global_max_concurrency": config.GLOBAL_MAX_CONCURRENCY,
        "global_max_qps": config.GLOBAL_MAX_QPS,
        # 代理模式
        "proxy_mode": config.PROXY_MODE,
        # 隧道代理配置
        "tunnel_proxy_url": config.TUNNEL_PROXY_URL,
        "tunnel_channels": config.TUNNEL_CHANNELS,
        "tunnel_rotate_interval": config.TUNNEL_ROTATE_INTERVAL,
        # DPS 优化参数
        "tunnel_max_concurrency": getattr(config, "TUNNEL_MAX_CONCURRENCY", 48),
        "tunnel_initial_concurrency": getattr(config, "TUNNEL_INITIAL_CONCURRENCY", 16),
        "per_channel_qps": getattr(config, "PER_CHANNEL_QPS", 3.0),
        "per_channel_max_concurrency": getattr(config, "PER_CHANNEL_MAX_CONCURRENCY", 12),
    }

_SETTINGS_FILE = os.path.join(config.BASE_DIR, "runtime_settings.json")

def _save_settings():
    """将当前设置持久化到 JSON 文件"""
    try:
        with open(_SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(_runtime_settings, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger.warning(f"⚠️ 设置保存失败: {e}")

def _load_settings() -> dict:
    """启动时加载持久化设置，不存在则返回默认值"""
    defaults = _default_settings()
    if not os.path.exists(_SETTINGS_FILE):
        return defaults
    try:
        with open(_SETTINGS_FILE, "r", encoding="utf-8") as f:
            saved = json.load(f)
        # 用保存的值覆盖默认值（兼容新增字段）
        for k in defaults:
            if k in saved:
                defaults[k] = saved[k]
        logger.info(f"⚙️ 已从 {_SETTINGS_FILE} 加载持久化设置")
        return defaults
    except Exception as e:
        logger.warning(f"⚠️ 加载持久化设置失败: {e}，使用默认值")
        return defaults

_runtime_settings = _load_settings()
# 设置版本号：每次修改 +1，Worker 比对版本号决定是否需要重载
_settings_version = 0

# ==================== 全局并发协调器 ====================
_global_coordinator = {
    "worker_metrics": {},       # {worker_id: {snapshot 字段 + reported_at}}
    "worker_quotas": {},        # {worker_id: {concurrency, qps, assigned_at}}
    "global_block_until": 0.0,  # monotonic 时间戳；> now 表示全局冷却中
    "block_count": 0,           # 累计全局封锁次数
    "recovery_epoch": 0,        # 每次全局封锁 +1，Worker 用于去重
    "recovery_jitter": {},      # {worker_id: float 0.0~1.0} 恢复抖动系数
}


def _allocate_quotas():
    """按健康度加权分配全局并发/QPS 预算给活跃 Worker"""
    now = time.time()
    active = {wid: info for wid, info in _worker_registry.items()
              if now - info["last_seen"] < 60}
    n = len(active)
    if n == 0:
        _global_coordinator["worker_quotas"] = {}
        return

    g = _global_coordinator
    total_conc = _runtime_settings["global_max_concurrency"]
    total_qps = _runtime_settings["global_max_qps"]

    # 全局冷却期内预算减半
    now_mono = time.monotonic()
    in_cooldown = now_mono < g["global_block_until"]
    if in_cooldown:
        total_conc = max(n * _runtime_settings["min_concurrency"], total_conc // 2)
        total_qps = max(n * 0.5, total_qps / 2)

    # 计算每个 Worker 的健康分（0.1 ~ 1.0）
    scores = {}
    for wid in active:
        metrics = g["worker_metrics"].get(wid, {})
        if not metrics or now - metrics.get("reported_at", 0) > 90:
            scores[wid] = 0.5  # 无数据或过期 → 默认分
        else:
            sr = metrics.get("success_rate", 0.95)
            br = metrics.get("block_rate", 0.0)
            # 封锁率惩罚权重 ×5：block_rate=10% → 分数减半
            score = sr * max(0.0, 1.0 - br * 5)
            scores[wid] = max(0.1, min(1.0, score))

    total_score = sum(scores.values())

    # 第一轮：按权重计算原始配额（不设下限，纯比例分配）
    raw_conc = {}
    raw_qps = {}
    for wid in active:
        weight = scores[wid] / total_score
        raw_conc[wid] = total_conc * weight
        raw_qps[wid] = total_qps * weight

        # 冷却期叠加抖动系数
        if in_cooldown:
            jitter = g["recovery_jitter"].get(wid, 0.5)
            raw_conc[wid] *= (0.5 + 0.5 * jitter)
            raw_qps[wid] *= (0.5 + 0.5 * jitter)

    # 第二轮：取整 + 下限保护，然后裁剪总量不超预算
    min_c = _runtime_settings["min_concurrency"]
    # 当 min_c × n > budget 时，预算优先：降低有效下限
    effective_min_c = min(min_c, max(1, int(total_conc / n)))
    int_conc = {wid: max(effective_min_c, int(v)) for wid, v in raw_conc.items()}

    # 如果总和超预算，按比例缩减（保留有效下限）
    sum_conc = sum(int_conc.values())
    if sum_conc > total_conc:
        excess = sum_conc - total_conc
        reducible = {wid: v - effective_min_c for wid, v in int_conc.items() if v > effective_min_c}
        total_reducible = sum(reducible.values())
        if total_reducible > 0:
            for wid in reducible:
                cut = int(excess * reducible[wid] / total_reducible)
                int_conc[wid] = max(effective_min_c, int_conc[wid] - cut)
        # 最终兜底：逐个削减直到不超预算
        while sum(int_conc.values()) > total_conc:
            for wid in sorted(int_conc, key=lambda w: int_conc[w], reverse=True):
                if int_conc[wid] > effective_min_c and sum(int_conc.values()) > total_conc:
                    int_conc[wid] -= 1

    sum_qps = sum(raw_qps.values())
    if sum_qps > total_qps and sum_qps > 0:
        scale = total_qps / sum_qps
        raw_qps = {wid: v * scale for wid, v in raw_qps.items()}

    # 单 Worker 并发上限：不超过 max_concurrency 设置
    per_worker_max = _runtime_settings["max_concurrency"]

    new_quotas = {}
    for wid in active:
        new_quotas[wid] = {
            "concurrency": min(int_conc[wid], per_worker_max),
            "qps": round(max(0.1, raw_qps[wid]), 2),
            "assigned_at": now,
        }

    g["worker_quotas"] = new_quotas


def _handle_worker_metrics(worker_id: str, metrics: dict):
    """处理 Worker 上报的 metrics，必要时触发全局封锁"""
    g = _global_coordinator
    now_mono = time.monotonic()

    # 存储 metrics
    g["worker_metrics"][worker_id] = {
        **metrics,
        "reported_at": time.time(),
    }

    # 检查是否触发全局封锁
    block_rate = metrics.get("block_rate", 0.0)
    threshold = _runtime_settings.get("block_rate_threshold", 0.05)

    if block_rate > threshold and now_mono >= g["global_block_until"]:
        cooldown = _runtime_settings.get("cooldown_after_block", 30)
        g["global_block_until"] = now_mono + cooldown
        g["block_triggered_by"] = worker_id
        g["block_count"] += 1
        g["recovery_epoch"] += 1

        # 为每个活跃 Worker 分配不同的恢复抖动系数
        for wid in _worker_registry:
            g["recovery_jitter"][wid] = random.random()

        logger.warning(
            f"⚠️ 全局封锁触发 by {worker_id} "
            f"(block_rate={block_rate:.1%}), "
            f"冷却 {cooldown}s, epoch={g['recovery_epoch']}"
        )

        # 立即重新分配配额（减半生效）
        _allocate_quotas()


async def _timeout_task_loop():
    """定期回退超时 processing 任务，并清理长期离线 Worker"""
    while True:
        try:
            db = await get_db()
            count = await db.reset_timeout_tasks()
            if count > 0:
                logger.info(f"回退了 {count} 个超时任务")
        except Exception as e:
            logger.error(f"超时任务回退异常: {e}")

        # 清理超过 10 分钟无心跳的 Worker（防止注册表无限增长）
        now = time.time()
        stale = [wid for wid, info in _worker_registry.items()
                 if now - info["last_seen"] > 600]
        for wid in stale:
            del _worker_registry[wid]
            # 同步清理协调器中的过期数据
            _global_coordinator["worker_metrics"].pop(wid, None)
            _global_coordinator["worker_quotas"].pop(wid, None)
            _global_coordinator["recovery_jitter"].pop(wid, None)
        if stale:
            logger.info(f"清理了 {len(stale)} 个长期离线 Worker")

        # 兜底配额重算（确保 Worker 离线后配额被回收）
        _allocate_quotas()

        await asyncio.sleep(60)


# ==================== API 端点 ====================

# --- 任务上传 ---
@app.post("/api/upload")
async def upload_asin_file(
    file: UploadFile = File(...),
    batch_name: str = Form(None),
    zip_code: str = Form(None),
    needs_screenshot: bool = Form(False),
):
    """
    上传 ASIN 文件（Excel/CSV）
    自动识别文件类型，提取 ASIN 列
    """
    db = await get_db()
    zip_code = zip_code or _runtime_settings["zip_code"]

    # 自动生成批次名
    if not batch_name:
        batch_name = f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    # 批次名白名单净化：只允许字母、数字、下划线、连字符
    batch_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', batch_name).strip('_')
    if not batch_name:
        batch_name = f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    # 读取文件内容
    content = await file.read()
    filename = (file.filename or "").lower()

    # 文件类型白名单
    allowed_extensions = ('.xlsx', '.xls', '.csv', '.txt')
    if not any(filename.endswith(ext) for ext in allowed_extensions):
        raise HTTPException(status_code=400, detail=f"不支持的文件类型，仅允许: {', '.join(allowed_extensions)}")

    asins = []
    try:
        if filename.endswith(('.xlsx', '.xls')):
            # Excel 文件
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        val = str(cell).strip().upper()
                        # ASIN 格式：10位字母数字，以 B 开头
                        if re.match(r'^B[0-9A-Z]{9}$', val):
                            asins.append(val)
            wb.close()
        elif filename.endswith('.csv'):
            # CSV 文件
            text = content.decode('utf-8-sig')
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                for cell in row:
                    val = cell.strip().upper()
                    if re.match(r'^B[0-9A-Z]{9}$', val):
                        asins.append(val)
        else:
            # 纯文本（每行一个 ASIN）
            text = content.decode('utf-8-sig')
            for line in text.splitlines():
                val = line.strip().upper()
                if re.match(r'^B[0-9A-Z]{9}$', val):
                    asins.append(val)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")

    if not asins:
        raise HTTPException(status_code=400, detail="文件中未找到有效的 ASIN")

    # 去重
    asins = list(dict.fromkeys(asins))

    # 创建任务
    inserted = await db.create_tasks(batch_name, asins, zip_code, needs_screenshot)

    return {
        "status": "ok",
        "batch_name": batch_name,
        "total_asins": len(asins),
        "inserted": inserted,
        "zip_code": zip_code,
        "needs_screenshot": needs_screenshot,
    }


# --- Worker 拉取任务 ---
@app.get("/api/tasks/pull")
async def pull_tasks(
    worker_id: str = Query(...),
    count: int = Query(10),
):
    """Worker 拉取待处理任务"""
    db = await get_db()
    _register_worker(worker_id)

    # 不支持截图的 Worker 只拉取不需要截图的任务
    worker_info = _worker_registry.get(worker_id, {})
    screenshot_only = None if worker_info.get("enable_screenshot", True) else False

    tasks = await db.pull_tasks(worker_id, max(1, min(count, 50)), needs_screenshot=screenshot_only)
    
    if worker_id in _worker_registry:
        _worker_registry[worker_id]["tasks_pulled"] += len(tasks)

    return {"tasks": tasks}


# --- Worker 释放任务（优先采集队列切换时归还旧任务）---
@app.post("/api/tasks/release")
async def release_tasks(request: Request):
    """Worker 归还未处理的任务，立即重置为 pending（避免等 5 分钟超时）"""
    db = await get_db()
    data = await request.json()
    task_ids = data.get("task_ids", [])
    if not task_ids:
        return {"status": "ok", "released": 0}
    count = await db.release_tasks(task_ids)
    logger.info(f"🔄 Worker 归还了 {count} 个任务")
    return {"status": "ok", "released": count}


# --- Worker 提交结果 ---
@app.post("/api/tasks/result")
async def submit_result(request: Request):
    """Worker 提交采集结果"""
    db = await get_db()
    data = await request.json()
    
    task_id = data.get("task_id")
    worker_id = data.get("worker_id", "unknown")
    success = data.get("success", False)
    result_data = data.get("result")

    _register_worker(worker_id)
    if worker_id in _worker_registry:
        _worker_registry[worker_id]["results_submitted"] += 1

    if success and result_data:
        # 保存结果
        await db.save_result(result_data)
        await db.mark_task_done(task_id, worker_id)
    else:
        await db.mark_task_failed(task_id, worker_id,
                                  error_type=data.get("error_type"),
                                  error_detail=data.get("error_detail"))

    return {"status": "ok"}


# --- Worker 批量提交结果 ---
@app.post("/api/tasks/result/batch")
async def submit_result_batch(request: Request):
    """Worker 批量提交采集结果（统一提交，减少磁盘 IO）

    修复：通过 db.batch_submit_results() 使用写锁序列化，
    防止与 pull_tasks 并发时出现 "cannot start a transaction within a transaction" 错误
    """
    db = await get_db()
    data = await request.json()
    results_list = data.get("results", [])

    # 更新 worker 注册信息（内存操作，不需要锁）
    for item in results_list:
        worker_id = item.get("worker_id", "unknown")
        _register_worker(worker_id)
        if worker_id in _worker_registry:
            _worker_registry[worker_id]["results_submitted"] += 1

    # 数据库操作委托给 Database 方法（内部使用写锁序列化）
    count = await db.batch_submit_results(results_list, RESULT_FIELDS)

    return {"status": "ok", "count": count}


# --- 进度查询 ---
@app.get("/api/progress/{batch_name}")
async def get_progress(batch_name: str):
    """获取指定批次的采集进度"""
    db = await get_db()
    progress = await db.get_progress(batch_name)
    return progress


@app.get("/api/progress")
async def get_overall_progress():
    """获取总体进度"""
    db = await get_db()
    progress = await db.get_progress()
    return progress


# --- 批次列表 ---
@app.get("/api/batches")
async def get_batches():
    """获取所有批次列表"""
    db = await get_db()
    batches = await db.get_batch_list()
    return {"batches": batches}


# --- 结果查询 ---
@app.get("/api/results")
async def get_results(
    batch_name: str = Query(None),
    page: int = Query(1),
    per_page: int = Query(50),
    search: str = Query(None),
):
    """分页获取采集结果"""
    db = await get_db()
    results, total = await db.get_results(batch_name, page, per_page, search)
    return {
        "results": results,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page,
    }


# --- 数据导出 ---
@app.get("/api/export/{batch_name}")
async def export_data(
    batch_name: str,
    format: str = Query("excel"),
):
    """导出采集数据（Excel/CSV）"""
    db = await get_db()
    results = await db.get_all_results(batch_name)

    if not results:
        raise HTTPException(status_code=404, detail="该批次无数据")

    if format == "csv":
        return _export_csv(results, batch_name)
    else:
        return _export_excel(results, batch_name)


@app.get("/api/export/{batch_name}/screenshots")
async def export_screenshots(batch_name: str):
    """批量下载某批次的所有截图（ZIP 打包，使用临时文件避免内存峰值）"""
    import tempfile
    # 防路径穿越：只允许安全字符
    safe_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', batch_name)
    screenshot_dir = os.path.realpath(os.path.join(config.STATIC_DIR, "screenshots", safe_name))
    screenshots_root = os.path.realpath(os.path.join(config.STATIC_DIR, "screenshots"))
    if not screenshot_dir.startswith(screenshots_root + os.sep):
        raise HTTPException(status_code=400, detail="非法批次名")
    if not os.path.isdir(screenshot_dir):
        raise HTTPException(status_code=404, detail="该批次无截图")

    png_files = [f for f in os.listdir(screenshot_dir) if f.endswith(".png")]
    if not png_files:
        raise HTTPException(status_code=404, detail="该批次无截图文件")

    # 写入临时文件，避免大批次截图撑爆内存
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".zip")
    try:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname in sorted(png_files):
                fpath = os.path.join(screenshot_dir, fname)
                zf.write(fpath, fname)
        tmp_path = tmp.name
        tmp.close()
    except Exception:
        tmp.close()
        os.unlink(tmp.name)
        raise

    async def _stream_and_cleanup():
        try:
            with open(tmp_path, "rb") as f:
                while True:
                    chunk = f.read(65536)
                    if not chunk:
                        break
                    yield chunk
        finally:
            os.unlink(tmp_path)

    return StreamingResponse(
        _stream_and_cleanup(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_screenshots.zip"'},
    )


def _parse_price(s: str) -> float | None:
    """解析 '$12.99' 格式的价格字符串为浮点数，失败返回 None。"""
    if not s or s == "N/A":
        return None
    s = s.strip().replace(",", "")
    if s.startswith("$"):
        s = s[1:]
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _prepare_export_rows(results: List[Dict]):
    """准备导出数据，返回 (headers, field_keys, rows)。"""
    field_keys = list(RESULT_FIELDS)
    headers = [config.HEADER_MAP.get(f, f) for f in field_keys]
    rows = [[str(row_data.get(f, "")) for f in field_keys] for row_data in results]

    # 在 "BuyBox 运费" 后插入 "总价" 列
    shipping_header = "BuyBox 运费"
    if shipping_header in headers:
        insert_idx = headers.index(shipping_header) + 1
        headers.insert(insert_idx, "总价")
        for i, row_data in enumerate(results):
            price = _parse_price(str(row_data.get("buybox_price", "")))
            shipping_str = str(row_data.get("buybox_shipping", ""))
            if price is None:
                total = "N/A"
            elif shipping_str.upper() == "FREE":
                total = f"${price:.2f}"
            else:
                shipping = _parse_price(shipping_str)
                if shipping is None:
                    total = f"${price:.2f}"
                else:
                    total = f"${price + shipping:.2f}"
            rows[i].insert(insert_idx, total)

    return headers, field_keys, rows


def _export_excel(results: List[Dict], batch_name: str):
    """导出 Excel 文件"""
    headers, _, rows = _prepare_export_rows(results)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采集结果"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={batch_name}.xlsx"},
    )


def _export_csv(results: List[Dict], batch_name: str):
    """导出 CSV 文件"""
    headers, _, rows = _prepare_export_rows(results)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)

    content = output.getvalue().encode('utf-8-sig')
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={batch_name}.csv"},
    )


# --- Worker 监控 ---
@app.get("/api/workers")
async def get_workers():
    """获取在线 worker 列表"""
    now = time.time()
    workers = []
    for wid, info in _worker_registry.items():
        elapsed = now - info["last_seen"]
        is_online = elapsed < 60
        # 在线：now - first_seen；离线：last_seen - first_seen（冻结时长）
        uptime = (now if is_online else info["last_seen"]) - info["first_seen"]
        quota = _global_coordinator["worker_quotas"].get(wid, {})
        metrics = _global_coordinator["worker_metrics"].get(wid, {})
        workers.append({
            "worker_id": wid,
            "status": "online" if is_online else "offline",
            "last_seen": datetime.fromtimestamp(info["last_seen"]).strftime("%H:%M:%S"),
            "tasks_pulled": info["tasks_pulled"],
            "results_submitted": info["results_submitted"],
            "uptime": int(uptime),
            "quota_concurrency": quota.get("concurrency"),
            "quota_qps": quota.get("qps"),
            "success_rate": round(metrics["success_rate"] * 100, 2) if "success_rate" in metrics else None,
            "block_rate": round(metrics["block_rate"] * 100, 2) if "block_rate" in metrics else None,
            "latency_p50": metrics.get("latency_p50"),
            "inflight": metrics.get("inflight"),
            "enable_screenshot": info.get("enable_screenshot", True),
            "ip": info.get("ip"),
        })
    return {"workers": workers}


@app.delete("/api/workers/{worker_id}")
async def remove_worker(worker_id: str):
    """移除单个离线 worker"""
    if worker_id in _worker_registry:
        del _worker_registry[worker_id]
        return {"status": "ok", "removed": worker_id}
    return {"status": "not_found"}


@app.delete("/api/workers")
async def remove_offline_workers():
    """清理所有离线 worker（超过 60 秒无心跳）"""
    now = time.time()
    offline = [wid for wid, info in _worker_registry.items() if now - info["last_seen"] >= 60]
    for wid in offline:
        del _worker_registry[wid]
    return {"status": "ok", "removed": len(offline)}


# --- 全局并发协调 ---
@app.post("/api/worker/sync")
async def worker_sync(request: Request):
    """
    Worker 综合同步端点（每 30s 调用一次）
    功能：心跳 + 上报 metrics + 拉取 settings + 接收配额
    """
    data = await request.json()
    worker_id = data.get("worker_id")
    if not worker_id:
        raise HTTPException(400, "worker_id required")

    metrics = data.get("metrics")

    # 获取客户端 IP
    client_ip = request.client.host if request.client else None

    _register_worker(
        worker_id,
        enable_screenshot=data.get("enable_screenshot"),
        ip=client_ip,
    )

    # 处理 metrics（可能触发全局封锁）
    if metrics:
        _handle_worker_metrics(worker_id, metrics)

    # 确保配额是最新的
    _allocate_quotas()

    # 构建响应
    g = _global_coordinator
    quota = g["worker_quotas"].get(worker_id, {})
    now_mono = time.monotonic()
    in_cooldown = now_mono < g["global_block_until"]

    return {
        **_runtime_settings,
        "_version": _settings_version,
        "_quota": {
            "concurrency": quota.get("concurrency", _runtime_settings["max_concurrency"]),
            "qps": quota.get("qps", _runtime_settings["token_bucket_rate"]),
        },
        "_global_block": {
            "active": in_cooldown,
            "remaining_s": max(0, int(g["global_block_until"] - now_mono)) if in_cooldown else 0,
            "triggered_by": g.get("block_triggered_by") if in_cooldown else None,
            "epoch": g["recovery_epoch"],
        },
        "_recovery_jitter": g["recovery_jitter"].get(worker_id, 0.5),
    }


@app.get("/api/coordinator")
async def get_coordinator_state():
    """全局并发协调状态（可观测性端点）"""
    g = _global_coordinator
    now_mono = time.monotonic()
    in_cooldown = now_mono < g["global_block_until"]

    per_worker = {}
    for wid in g["worker_quotas"]:
        quota = g["worker_quotas"].get(wid, {})
        metrics = g["worker_metrics"].get(wid, {})
        per_worker[wid] = {
            "quota_concurrency": quota.get("concurrency"),
            "quota_qps": quota.get("qps"),
            "success_rate": round(metrics["success_rate"] * 100, 2) if "success_rate" in metrics else None,
            "block_rate": round(metrics["block_rate"] * 100, 2) if "block_rate" in metrics else None,
            "latency_p50": metrics.get("latency_p50"),
            "inflight": metrics.get("inflight"),
            "current_concurrency": metrics.get("current_concurrency"),
        }

    return {
        "global_max_concurrency": _runtime_settings["global_max_concurrency"],
        "global_max_qps": _runtime_settings["global_max_qps"],
        "allocated_concurrency": sum(
            q.get("concurrency", 0) for q in g["worker_quotas"].values()
        ),
        "allocated_qps": round(sum(
            q.get("qps", 0) for q in g["worker_quotas"].values()
        ), 2),
        "active_workers": len(g["worker_quotas"]),
        "global_block": {
            "active": in_cooldown,
            "remaining_s": max(0, int(g["global_block_until"] - now_mono)) if in_cooldown else 0,
            "block_count": g["block_count"],
            "recovery_epoch": g["recovery_epoch"],
        },
        "per_worker": per_worker,
    }


# --- 诊断端点 ---
_server_start_time = time.time()

@app.get("/api/diagnostic")
async def diagnostic():
    """
    系统诊断端点：用于远程检查 server/worker 运行状态、配置参数、代码版本

    用途：
    - 确认 DMIT 是否部署了最新代码（通过 code_version 中的关键参数判断）
    - 监控 worker 活跃度和任务吞吐量
    - 检测数据库健康状态
    """
    db = await get_db()
    progress = await db.get_progress()

    # Worker 活跃度
    now = time.time()
    active_workers = []
    for wid, info in _worker_registry.items():
        age = now - info["last_seen"]
        active_workers.append({
            "worker_id": wid,
            "alive": age < 60,
            "last_seen_ago_s": round(age, 1),
            "tasks_pulled": info["tasks_pulled"],
            "results_submitted": info["results_submitted"],
        })

    # 数据库写锁状态
    db_lock_status = "locked" if db._write_lock.locked() else "free"

    return {
        "server_uptime_s": round(now - _server_start_time, 1),
        "server_uptime_h": round((now - _server_start_time) / 3600, 2),
        "task_progress": progress,
        "workers": active_workers,
        "db_write_lock": db_lock_status,
        "code_version": {
            # 这些关键参数可以判断是否部署了最新优化代码
            "token_bucket_rate": config.TOKEN_BUCKET_RATE,
            "initial_concurrency": config.INITIAL_CONCURRENCY,
            "max_concurrency": config.MAX_CONCURRENCY,
            "target_latency_s": config.TARGET_LATENCY_S,
            "max_latency_s": config.MAX_LATENCY_S,
            "cooldown_after_block_s": 15,
            "proxy_mode": config.PROXY_MODE,
            "has_write_lock": hasattr(db, '_write_lock'),  # True = 新版代码
        },
        "runtime_settings_version": _settings_version,
    }


# --- 设置管理 ---
@app.get("/api/settings")
async def get_settings():
    """获取当前运行时设置（含版本号，Worker 用于增量同步）"""
    return {**_runtime_settings, "_version": _settings_version}


def _normalize_proxy_url(raw: str) -> str:
    """自动识别多种代理 URL 格式，统一转换为 http://user:pwd@host:port"""
    raw = raw.strip()
    if not raw:
        return raw

    # 已有 scheme → 直接返回
    if re.match(r'^https?://', raw) or re.match(r'^socks[45h]?://', raw):
        return raw

    # 空格分隔: "host:port user:pwd" 等
    if ' ' in raw:
        parts = raw.split(None, 1)
        if len(parts) == 2:
            left, right = parts[0], parts[1]
            right_segs = right.split(':')
            if ':' in left and ':' in right:
                return f"http://{right}@{left}"
            elif ':' not in left and len(right_segs) == 3:
                port, user, pwd = right_segs
                return f"http://{user}:{pwd}@{left}:{port}"
            elif ':' not in left and len(right_segs) == 2:
                return f"http://{right}@{left}"

    # @ 分隔: host:port@user:pwd 或 user:pwd@host:port
    if '@' in raw:
        left, right = raw.split('@', 1)
        left_parts = left.split(':')
        right_parts = right.split(':')
        if len(left_parts) == 2 and len(right_parts) == 2:
            left_has_dot = '.' in left_parts[0]
            right_has_dot = '.' in right_parts[0]
            if left_has_dot and not right_has_dot:
                return f"http://{right}@{left}"
            elif right_has_dot and not left_has_dot:
                return f"http://{left}@{right}"
            elif left_parts[1].isdigit() and not right_parts[1].isdigit():
                return f"http://{right}@{left}"
            else:
                return f"http://{left}@{right}"

    # 四段冒号分隔: host:port:user:pwd 或 user:pwd:host:port
    parts = raw.split(':')
    if len(parts) == 4:
        if '.' in parts[0]:
            host, port, user, pwd = parts
        else:
            user, pwd, host, port = parts
        return f"http://{user}:{pwd}@{host}:{port}"

    return raw


@app.put("/api/settings")
async def update_settings(request: Request):
    """更新运行时设置（含类型与范围校验）"""
    global _settings_version
    data = await request.json()

    # 类型与范围校验规则：(type, min, max)
    _validators = {
        "zip_code":             (str,   None, None),
        "max_retries":          (int,   1,    10),
        "request_timeout":      (int,   5,    60),
        "proxy_api_url":        (str,   None, None),
        "token_bucket_rate":    (float, 0.5,  50),
        "initial_concurrency":  (int,   1,    50),
        "min_concurrency":      (int,   1,    20),
        "max_concurrency":      (int,   2,    500),
        "session_rotate_every": (int,   50,   10000),
        "screenshot_browsers":    (int, 1,    10),
        "screenshot_pages_per_browser": (int, 1, 10),
        "adjust_interval":      (int,   3,    60),
        "target_latency":       (float, 1,    30),
        "max_latency":          (float, 2,    60),
        "target_success_rate":  (float, 0.5,  1),
        "min_success_rate":     (float, 0.3,  1),
        "block_rate_threshold": (float, 0.01, 0.5),
        "cooldown_after_block": (int,   5,    120),
        "proxy_bandwidth_mbps": (int,   0,    100),
        "global_max_concurrency": (int,  2,    500),
        "global_max_qps":         (float, 0.5, 100),
        # 代理模式
        "proxy_mode":           (str,   None, None),
        # 隧道代理配置
        "tunnel_proxy_url":     (str,   None, None),
        "tunnel_channels":      (int,   1,    32),
        "tunnel_rotate_interval": (int, 10,   300),
        # DPS 优化参数
        "tunnel_max_concurrency":     (int,   8,    200),
        "tunnel_initial_concurrency": (int,   2,    100),
        "per_channel_qps":            (float, 0.5,  20),
        "per_channel_max_concurrency": (int,   1,    30),
    }

    changed = False
    errors = []
    old_values = {}  # 保存旧值用于交叉校验失败回滚

    # 代理地址：提交前先规范化，确保比较用规范格式
    if "tunnel_proxy_url" in data and data["tunnel_proxy_url"]:
        data["tunnel_proxy_url"] = _normalize_proxy_url(data["tunnel_proxy_url"])

    for key in _runtime_settings:
        if key not in data or data[key] == _runtime_settings[key]:
            continue
        val = data[key]
        validator = _validators.get(key)
        if not validator:
            continue

        expected_type, lo, hi = validator
        # 类型转换与校验
        try:
            if expected_type is int:
                val = int(val)
            elif expected_type is float:
                val = float(val)
            elif expected_type is str:
                val = str(val)
        except (ValueError, TypeError):
            errors.append(f"{key}: 期望 {expected_type.__name__}，收到 {type(val).__name__}")
            continue

        # 代理地址自动识别：快代理 API 地址 → 存入 proxy_api_url；否则当固定代理处理
        if key == "tunnel_proxy_url" and val:
            if "kdlapi.com" in val or "kdl.cc" in val:
                # 快代理 API 地址 → 存到 proxy_api_url，清空 tunnel_proxy_url
                old_values.setdefault("proxy_api_url", _runtime_settings["proxy_api_url"])
                _runtime_settings["proxy_api_url"] = val
                old_values.setdefault("tunnel_proxy_url", _runtime_settings["tunnel_proxy_url"])
                _runtime_settings["tunnel_proxy_url"] = ""
                logger.info(f"⚙️ 检测到快代理 API 地址，已存入 proxy_api_url")
                changed = True
                continue
            original = val
            val = _normalize_proxy_url(val)
            if val != original:
                logger.info(f"⚙️ 代理 URL 自动转换: {original} → {val}")

        # 范围校验
        if lo is not None and val < lo:
            errors.append(f"{key}: 不能小于 {lo}")
            continue
        if hi is not None and val > hi:
            errors.append(f"{key}: 不能大于 {hi}")
            continue

        old_values[key] = _runtime_settings[key]
        _runtime_settings[key] = val
        changed = True

    if errors:
        # 单字段校验失败，回滚已写入的值
        for k, v in old_values.items():
            _runtime_settings[k] = v
        return JSONResponse(
            status_code=422,
            content={"status": "error", "errors": errors,
                     "settings": _runtime_settings, "_version": _settings_version}
        )

    # proxy_mode 枚举校验
    if "proxy_mode" in data and _runtime_settings.get("proxy_mode") not in ("tps", "tunnel"):
        for k, v in old_values.items():
            _runtime_settings[k] = v
        return JSONResponse(
            status_code=422,
            content={"status": "error",
                     "errors": [f"proxy_mode 必须为 'tps' 或 'tunnel'，收到 '{_runtime_settings.get('proxy_mode')}'"],
                     "settings": _runtime_settings, "_version": _settings_version}
        )

    # 代理模式切换联动：自动应用预设参数（仅用户未显式指定的字段）
    _MODE_PRESETS = {
        "tps": {
            "token_bucket_rate":    5.0,
            "initial_concurrency":  8,
            "min_concurrency":      4,
            "cooldown_after_block": 15,
            "proxy_bandwidth_mbps": 0,
            "global_max_qps":       5.0,
            "request_timeout":      15,
        },
        "tunnel": {
            "token_bucket_rate":    15.0,
            "initial_concurrency":  28,
            "min_concurrency":      12,
            "max_concurrency":      80,     # 允许系统探索更高并发来填满带宽管道
            "cooldown_after_block": 60,
            "proxy_bandwidth_mbps": 0,      # 禁用带宽限制（解压后字节率 ≠ 线路带宽）
            "global_max_qps":       15.0,
            "request_timeout":      45,     # 大页面(1.85MB)在高并发下下载慢，给够时间
            "target_latency":       25.0,   # 高并发共享带宽 → 延迟=C×pageSize/BW，允许到25s再停止增速
            "max_latency":          35.0,   # 仅在接近超时(45s)时才减速，避免误判带宽延迟
            "min_success_rate":     0.75,   # IP 轮换时成功率暂降，容忍到 75%
        },
    }
    if "proxy_mode" in old_values:
        # proxy_mode 确实发生了变化，应用预设
        new_mode = _runtime_settings["proxy_mode"]
        preset = _MODE_PRESETS.get(new_mode, {})
        linked = []
        for pk, pv in preset.items():
            # 只覆盖用户本次请求中没有显式提交的字段
            if pk not in data:
                old_values.setdefault(pk, _runtime_settings[pk])
                _runtime_settings[pk] = pv
                linked.append(f"{pk}={pv}")
                changed = True
        if linked:
            logger.info(f"⚙️ 模式切换联动 ({new_mode}): {', '.join(linked)}")

    # 交叉约束校验（所有单字段校验通过后）
    cross_errors = []
    s = _runtime_settings
    if s["min_concurrency"] > s["max_concurrency"]:
        cross_errors.append(f"min_concurrency({s['min_concurrency']}) 不能大于 max_concurrency({s['max_concurrency']})")
    if s["initial_concurrency"] > s["max_concurrency"]:
        cross_errors.append(f"initial_concurrency({s['initial_concurrency']}) 不能大于 max_concurrency({s['max_concurrency']})")
    if s["initial_concurrency"] < s["min_concurrency"]:
        cross_errors.append(f"initial_concurrency({s['initial_concurrency']}) 不能小于 min_concurrency({s['min_concurrency']})")
    if s["target_latency"] >= s["max_latency"]:
        cross_errors.append(f"target_latency({s['target_latency']}) 必须小于 max_latency({s['max_latency']})")
    if s["min_success_rate"] > s["target_success_rate"]:
        cross_errors.append(f"min_success_rate({s['min_success_rate']}) 不能大于 target_success_rate({s['target_success_rate']})")
    if cross_errors:
        # 交叉校验失败，回滚所有已写入的值
        for k, v in old_values.items():
            _runtime_settings[k] = v
        return JSONResponse(
            status_code=422,
            content={"status": "error", "errors": cross_errors,
                     "settings": _runtime_settings, "_version": _settings_version}
        )

    if changed:
        _settings_version += 1
        _save_settings()
        logger.info(f"⚙️ 设置已更新并持久化 (version={_settings_version})")
    return {"status": "ok", "settings": _runtime_settings, "_version": _settings_version}


@app.post("/api/settings/reset")
async def reset_settings():
    """恢复所有设置为默认值"""
    global _settings_version
    _runtime_settings.clear()
    _runtime_settings.update(_default_settings())
    _settings_version += 1
    _save_settings()
    logger.info(f"⚙️ 设置已恢复默认并持久化 (version={_settings_version})")
    return {"status": "ok", "settings": _runtime_settings, "_version": _settings_version}


# --- 批次操作 ---
@app.get("/api/batches/{batch_name}/errors")
async def batch_errors(batch_name: str):
    """获取批次的错误分类统计和失败任务详情"""
    db = await get_db()
    summary = await db.get_error_summary(batch_name)
    failed_tasks = await db.get_failed_tasks(batch_name)
    return {"summary": summary, "tasks": failed_tasks}


@app.post("/api/batches/{batch_name}/retry")
async def retry_batch(batch_name: str):
    """重试批次中所有失败的任务"""
    db = await get_db()
    await db.retry_all_failed(batch_name)
    return {"status": "ok"}


@app.post("/api/batches/{batch_name}/prioritize")
async def prioritize_batch(batch_name: str):
    """将批次中所有 pending 任务设为高优先级"""
    db = await get_db()
    count = await db.prioritize_batch(batch_name, priority=10)
    logger.info(f"🚀 批次 {batch_name} 已设为优先采集 ({count} 个任务)")
    return {"status": "ok", "updated": count}


@app.delete("/api/batches/{batch_name}")
async def delete_batch(batch_name: str):
    """删除批次（含数据库记录 + 截图文件）"""
    db = await get_db()
    await db.delete_batch(batch_name)

    # 清理该批次的截图目录
    safe_batch = re.sub(r'[^a-zA-Z0-9_\-]', '_', batch_name)
    screenshot_dir = os.path.join(config.STATIC_DIR, "screenshots", safe_batch)
    if os.path.isdir(screenshot_dir):
        import shutil
        shutil.rmtree(screenshot_dir, ignore_errors=True)
        logger.info(f"🗑️ 已清理截图目录: {safe_batch}")

    return {"status": "ok"}


@app.delete("/api/database")
async def clear_database():
    """清空数据库中所有数据（tasks + results + 截图文件）"""
    db = await get_db()
    counts = await db.clear_all()

    # 清理整个截图目录
    screenshots_root = os.path.join(config.STATIC_DIR, "screenshots")
    if os.path.isdir(screenshots_root):
        import shutil
        shutil.rmtree(screenshots_root, ignore_errors=True)
        os.makedirs(screenshots_root, exist_ok=True)  # 重建空目录
        logger.info("🗑️ 已清理所有截图文件")

    return {"status": "ok", **counts}


# --- 截图上传 ---
@app.post("/api/tasks/screenshot")
async def upload_screenshot(
    file: UploadFile = File(...),
    batch_name: str = Form(...),
    asin: str = Form(...),
):
    """Worker 上传截图文件，保存到 static/screenshots/ 并更新 results 表"""
    db = await get_db()

    # 净化路径，防路径穿越
    safe_batch = re.sub(r'[^a-zA-Z0-9_\-]', '_', batch_name)
    safe_asin = re.sub(r'[^A-Z0-9]', '', asin.upper())
    screenshot_dir = os.path.join(config.STATIC_DIR, "screenshots", safe_batch)
    os.makedirs(screenshot_dir, exist_ok=True)

    filename = f"{safe_asin}.png"
    filepath = os.path.join(screenshot_dir, filename)
    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    rel_path = f"/static/screenshots/{safe_batch}/{filename}"
    await db.update_screenshot_path(batch_name, asin, rel_path)

    logger.info(f"📸 截图已保存: {batch_name}/{asin} ({len(content)} bytes)")
    return {"status": "ok", "path": rel_path}


# --- Worker 下载包 ---
_WORKER_FILES = [
    "worker.py", "config.py", "proxy.py", "session.py",
    "parser.py", "metrics.py", "adaptive.py", "models.py",
    "requirements-worker.txt",
]


@app.get("/api/worker/download")
async def download_worker(request: Request, mode: str = "full"):
    """打包 Worker 所需文件为 ZIP 下载

    mode=full  : 完整安装包（源码 + requirements + 启动脚本），首次安装用
    mode=update: 仅 Python 源码文件，已安装过的机器用于更新代码
    """
    # --- mode=update: 仅打包 .py 源码，直接解压覆盖即可 ---
    if mode == "update":
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for fname in _WORKER_FILES:
                if not fname.endswith(".py"):
                    continue
                fpath = os.path.join(config.BASE_DIR, fname)
                if os.path.exists(fpath):
                    zf.write(fpath, fname)  # 无子目录，直接覆盖
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition": "attachment; filename=worker-update.zip"},
        )

    # --- mode=full: 完整安装包 ---
    host = request.headers.get("host", f"127.0.0.1:{config.SERVER_PORT}")
    scheme = "https" if request.headers.get("x-forwarded-proto") == "https" else "http"
    server_url = f"{scheme}://{host}"

    # 国内镜像地址
    pypi_mirror = "https://pypi.tuna.tsinghua.edu.cn/simple/"
    playwright_mirror = "https://registry.npmmirror.com/-/binary/playwright"

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Python 源码 + requirements
        for fname in _WORKER_FILES:
            fpath = os.path.join(config.BASE_DIR, fname)
            if os.path.exists(fpath):
                zf.write(fpath, f"worker/{fname}")

        # 可选截图依赖
        zf.writestr("worker/requirements-screenshot.txt",
                     "# 截图功能（可选），安装后需运行: playwright install chromium\n"
                     "playwright>=1.40.0\n")

        # 启动脚本 - macOS/Linux
        start_sh = f'''#!/bin/bash
# Amazon Scraper v2 - Worker 启动脚本
# 服务器地址: {server_url}

set -e
cd "$(dirname "$0")"
SERVER="{server_url}"

# 国内镜像加速
PYPI_MIRROR="{pypi_mirror}"
export PLAYWRIGHT_DOWNLOAD_HOST="{playwright_mirror}"

# 检测 Python
if command -v python3 &>/dev/null; then
    PY=python3
elif command -v python &>/dev/null; then
    PY=python
else
    echo "错误: 未找到 Python，请先安装 Python 3.10+"
    exit 1
fi

echo "使用 Python: $($PY --version)"

# 创建虚拟环境（首次运行）
if [ ! -d ".venv" ]; then
    echo "创建虚拟环境..."
    $PY -m venv .venv
fi

# 激活虚拟环境
source .venv/bin/activate

# 安装依赖（首次运行或依赖更新）
if [ ! -f ".deps_installed" ] || [ "requirements-worker.txt" -nt ".deps_installed" ]; then
    echo "安装核心依赖（使用清华镜像加速）..."
    pip install -q -i "$PYPI_MIRROR" -r requirements-worker.txt
    echo "安装截图组件 (playwright)..."
    pip install -q -i "$PYPI_MIRROR" -r requirements-screenshot.txt
    echo "下载 Chromium 浏览器（使用国内镜像加速）..."
    playwright install chromium
    touch .deps_installed
fi

echo ""
echo "========================================="
echo "  Amazon Scraper v2 - Worker"
echo "  服务器: $SERVER"
echo "========================================="
echo ""

python worker.py --server "$SERVER" "$@"
'''
        zf.writestr("worker/start.sh", start_sh)

        # 启动脚本 - Windows（CRLF 行尾 + 纯 ASCII，兼容所有 Windows 区域设置）
        start_bat_lines = [
            '@echo off',
            'chcp 65001 >nul 2>&1',
            'title Amazon Scraper v2 - Worker',
            'cd /d "%~dp0"',
            f'set SERVER={server_url}',
            '',
            'REM Mirror acceleration for China',
            f'set PYPI_MIRROR={pypi_mirror}',
            f'set PLAYWRIGHT_DOWNLOAD_HOST={playwright_mirror}',
            '',
            'REM Check Python installation',
            'python --version >nul 2>&1',
            'if errorlevel 1 (',
            '    echo [ERROR] Python not found. Please install Python 3.10+ first.',
            '    echo Download: https://www.python.org/downloads/',
            '    echo Make sure to check "Add Python to PATH" during installation.',
            '    pause',
            '    exit /b 1',
            ')',
            '',
            'echo Python found: ',
            'python --version',
            '',
            'if not exist ".venv" (',
            '    echo Creating virtual environment...',
            '    python -m venv .venv',
            ')',
            '',
            'call .venv\\Scripts\\activate.bat',
            '',
            'if not exist ".deps_installed" (',
            '    echo Installing dependencies - using Tsinghua PyPI mirror...',
            '    pip install -q -i %PYPI_MIRROR% -r requirements-worker.txt',
            '    echo Installing screenshot component - playwright...',
            '    pip install -q -i %PYPI_MIRROR% -r requirements-screenshot.txt',
            '    echo Downloading Chromium browser - using npmmirror...',
            '    playwright install chromium',
            '    echo. > .deps_installed',
            ')',
            '',
            'echo.',
            'echo =========================================',
            'echo   Amazon Scraper v2 - Worker',
            'echo   Server: %SERVER%',
            'echo =========================================',
            'echo.',
            '',
            'python worker.py --server "%SERVER%" %*',
            'pause',
        ]
        start_bat = '\r\n'.join(start_bat_lines) + '\r\n'
        zf.writestr("worker/start.bat", start_bat.encode('ascii'))

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=worker.zip"},
    )


# ==================== Web UI 路由 ====================

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    """仪表盘首页"""
    db = await get_db()
    progress = await db.get_progress()
    batches = await db.get_batch_list()
    
    # 计算速度（最近5分钟的完成数）
    now = time.time()
    active_workers = sum(1 for w in _worker_registry.values() if now - w["last_seen"] < 60)

    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "progress": progress,
        "batches": batches[:5],  # 最近5个批次
        "active_workers": active_workers,
        "total_workers": len(_worker_registry),
    })


@app.get("/tasks", response_class=HTMLResponse)
async def tasks_page(request: Request):
    """任务管理页面"""
    db = await get_db()
    batches = await db.get_batch_list()
    return templates.TemplateResponse("tasks.html", {
        "request": request,
        "batches": batches,
        "default_zip_code": _runtime_settings["zip_code"],
    })


@app.get("/results", response_class=HTMLResponse)
async def results_page(
    request: Request,
    batch_name: str = None,
    page: int = 1,
    search: str = None,
):
    """结果浏览页面"""
    db = await get_db()
    batches = await db.get_batch_list()
    results, total = await db.get_results(batch_name, page, 50, search)
    total_pages = (total + 49) // 50
    progress = await db.get_progress(batch_name)

    return templates.TemplateResponse("results.html", {
        "request": request,
        "results": results,
        "batches": batches,
        "current_batch": batch_name,
        "current_page": page,
        "total": total,
        "total_pages": total_pages,
        "search": search or "",
        "progress": progress,
    })


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    """设置页面"""
    return templates.TemplateResponse("settings.html", {
        "request": request,
        "settings": _runtime_settings,
        "config": {
            "port": config.SERVER_PORT,
            "timeout": config.REQUEST_TIMEOUT,
        },
    })


@app.get("/workers", response_class=HTMLResponse)
async def workers_page(request: Request):
    """Worker 监控页面"""
    now = time.time()
    workers = []
    for wid, info in _worker_registry.items():
        elapsed = now - info["last_seen"]
        is_online = elapsed < 60
        uptime = (now if is_online else info["last_seen"]) - info["first_seen"]
        workers.append({
            "worker_id": wid,
            "status": "online" if is_online else "offline",
            "last_seen": datetime.fromtimestamp(info["last_seen"]).strftime("%Y-%m-%d %H:%M:%S"),
            "tasks_pulled": info["tasks_pulled"],
            "results_submitted": info["results_submitted"],
            "uptime_min": int(uptime / 60),
        })
    return templates.TemplateResponse("workers.html", {
        "request": request,
        "workers": workers,
    })


# ==================== 主入口 ====================

if __name__ == "__main__":
    uvicorn.run(
        "server:app",
        host=config.SERVER_HOST,
        port=config.SERVER_PORT,
        reload=False,
        log_level="info",
    )
