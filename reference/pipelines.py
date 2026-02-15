import os
import time
import logging
import pandas as pd
import datetime
from openpyxl import load_workbook
from scrapy.exceptions import DropItem


class AmazonSpiderPipeline:
    def __init__(self):
        # --- [配置] 保存目录 ---
        self.base_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "产品采集数据")
        if not os.path.exists(self.base_dir):
            os.makedirs(self.base_dir)

        # --- 配置参数 ---
        self.batch_size = 100        # 每 100 条写入一次
        self.max_lines_per_file = 50000  # 单个 Excel 最大行数
        self.current_file_index = 1  # 当前分卷序号
        self.current_file_lines = 0  # 当前文件已写入行数

        # --- 生成本次运行的基础文件名 ---
        self.timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        self._update_excel_path() # 初始化第一个文件路径

        # --- 生成本次运行的源码文件夹 ---
        self.html_dir = os.path.join(self.base_dir, f"{self.timestamp}_source_html")
        if not os.path.exists(self.html_dir):
            os.makedirs(self.html_dir)

        self.data_buffer = []
        self.seen_asins = set()
        
        # --- 实时速度统计 ---
        self.total_items_written = 0
        self.last_flush_time = time.time()
        self.start_time = time.time()
        
        self.logger = logging.getLogger('Pipeline')

        print(f"✅ [初始化] 数据文件: {self.excel_path}")

        # 表头映射
        self.header_map = {
            'crawl_time': '商品采集时间', 'zip_code': '配送邮编', 'product_url': '产品链接',
            'asin': 'ASIN (商品ID)', 'title': '商品标题',
            'original_price': '商品原价', 'current_price': '当前价格', 'buybox_price': 'BuyBox 价格',
            'buybox_shipping': 'BuyBox 运费', 'is_fba': '是否 FBA 发货', 'stock_count': '库存数量',
            'stock_status': '库存状态', 'delivery_date': '配送到达时间', 'delivery_time': '配送时长',
            'brand': '品牌', 'model_number': '产品型号', 'country_of_origin': '原产国',
            'is_customized': '是否为定制产品', 'best_sellers_rank': '畅销排名', 'upc_list': 'UPC 列表',
            'ean_list': 'EAN 列表', 'package_dimensions': '包装尺寸', 'package_weight': '包装重量',
            'item_dimensions': '商品本体尺寸', 'item_weight': '商品本体重量', 'parent_asin': '父体 ASIN',
            'variation_asins': '变体 ASIN 列表', 'root_category_id': '根类目 ID', 'category_tree': '类目路径树',
            'bullet_points': '五点描述', 'image_urls': '商品图片链接', 'site': '站点',
            'manufacturer': '制造商', 'part_number': '部件编号', 'first_available_date': '上架时间',
            'long_description': '长描述', 'product_type': '商品类型', 'category_ids': '类目 ID 链'
        }

        # 列顺序
        self.column_order = [
            '商品采集时间', '配送邮编', '产品链接', 'ASIN (商品ID)', '商品标题', '商品原价', '当前价格',
            'BuyBox 价格', 'BuyBox 运费', '是否 FBA 发货', '库存数量', '库存状态',
            '配送到达时间', '配送时长', '品牌', '产品型号', '原产国', '是否为定制产品',
            '畅销排名', 'UPC 列表', 'EAN 列表', '包装尺寸', '包装重量',
            '商品本体尺寸', '商品本体重量', '父体 ASIN', '变体 ASIN 列表', '根类目 ID',
            '类目路径树', '五点描述', '商品图片链接', '站点', '制造商', '部件编号',
            '上架时间', '长描述', '商品类型', '类目 ID 链'
        ]

    def _update_excel_path(self):
        """更新当前 Excel 文件路径"""
        self.excel_path = os.path.join(self.base_dir, f"{self.timestamp}_产品采集_part{self.current_file_index}.xlsx")
        self.current_file_lines = 0

    def process_item(self, item, spider):
        asin = item.get('asin', 'UNKNOWN')
        title = item.get('title')

        # 1. 有效性检查
        if not title or title == 'N/A':
            raise DropItem(f"❌ [无效数据] 丢弃 ASIN: {asin}")

        # 2. ASIN 去重
        if asin in self.seen_asins:
            raise DropItem(f"♻️ [重复数据] 丢弃 ASIN: {asin}")
        if asin: 
            self.seen_asins.add(asin)

        # 3. 保存 HTML 源码 (根据 settings 开关)
        # 默认为 False，只有 Explicitly Set True 才保存
        save_html = spider.settings.getbool('SAVE_HTML_SOURCE', False)
        
        if save_html and 'source_html' in item:
            try:
                html_content = item.pop('source_html')
                if html_content:
                    safe_asin = "".join([c for c in asin if c.isalnum() or c in (' ', '.', '_')]).strip()
                    file_name = os.path.join(self.html_dir, f"{safe_asin}.html")
                    with open(file_name, 'w', encoding='utf-8') as f:
                        f.write(html_content)
            except Exception:
                pass
        else:
            # 如果不保存，也得把字段删了，免得写入 Excel 报错或占内存
            item.pop('source_html', None)

        # 4. 加入缓冲区
        self.data_buffer.append(dict(item))

        # 5. 落盘检测
        if len(self.data_buffer) >= self.batch_size:
            self.flush(spider)

        return item

    def close_spider(self, spider):
        """爬虫结束时，写入剩余数据"""
        self.flush(spider, force=True)
        
        # 计算总耗时和平均速度
        total_time = time.time() - self.start_time
        avg_speed = self.total_items_written / total_time * 60 if total_time > 0 else 0
        
        print(f"\n🎉 [采集全部完成]")
        print(f"📊 总计采集: {self.total_items_written} 条 | 平均速度: {avg_speed:.1f} 条/min")
        print(f"📂 Excel文件: {self.excel_path} (最新卷)")
        if spider.settings.getbool('SAVE_HTML_SOURCE', False):
            print(f"📂 源码文件夹: {self.html_dir}\n")

    def flush(self, spider=None, force=False):
        if not self.data_buffer:
            return

        # --- 检查是否需要分卷 ---
        # 如果当前文件已满（且不是第一次写），则切换文件
        if self.current_file_lines >= self.max_lines_per_file:
            self.current_file_index += 1
            self._update_excel_path()
            print(f"📂 [自动分卷] 达到 {self.max_lines_per_file} 条，切换到新文件: {self.excel_path}")

        current_time = time.time()
        items_count = len(self.data_buffer)

        df = pd.DataFrame(self.data_buffer)
        df_export = df.rename(columns=self.header_map)

        final_cols = [c for c in self.column_order if c in df_export.columns]
        final_cols += [c for c in df_export.columns if c not in final_cols]
        df_export = df_export[final_cols]

        try:
            self._append_to_excel(df_export, self.excel_path)
            
            # 更新统计
            self.total_items_written += items_count
            self.current_file_lines += items_count # 更新当前文件行数
            
            # 计算速度
            time_diff = current_time - self.last_flush_time
            speed = items_count / time_diff * 60 if time_diff > 0 else 0
            
            # 新日志格式
            self.logger.info(
                f"✅ [写入] 本次 {items_count} 条 | "
                f"卷{self.current_file_index} 已存 {self.current_file_lines} | "
                f"总计 {self.total_items_written} | "
                f"速度 {speed:.1f}/min"
            )
            
            self.last_flush_time = current_time
            
        except Exception as e:
            self.logger.error(f"❌ [写入错误] {e}")

        self.data_buffer.clear()

    def _append_to_excel(self, df_export: pd.DataFrame, file_path: str):
        # 1. 首次写入
        if not os.path.exists(file_path):
            df_export.to_excel(file_path, index=False)
            return

        # 2. 追加写入
        try:
            book = load_workbook(file_path)
            start_row = book.active.max_row if book.active else 0
            book.close()
        except Exception:
            start_row = 0

        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df_export.to_excel(writer, index=False, header=False, startrow=start_row)